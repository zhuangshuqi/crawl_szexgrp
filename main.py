import os
from datetime import datetime
import requests
import json
from tenacity import retry, stop_after_attempt, wait_exponential
from openpyxl import Workbook
from openpyxl.styles import Font
import argparse

class Jyfwxx:
    BASE_URL = "https://www.szexgrp.com"
    RENT_URL = "https://rent.szexgrp.com"

    def __init__(self, title="城市更新", size=50, dir_path=r"D:\work\交易信息\附件"):
        self.url = f"{self.BASE_URL}/cms/api/v1/trade/content/page"
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36",
            "Content-Type": "application/json",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Origin": self.BASE_URL,
            "Referer": f"{self.BASE_URL}/",
            "Connection": "keep-alive"
        }
        self.params = {"channelId":2855,
                        "fields":[],
                        "title":title,
                        "page":0,
                        "size":size,
                        "parentBusinessType":"",
                        "modelId":1378,
                        "siteId":1}
        self.totalPages = 0
        self.totalElements = 0
        self.content_list = []
        self.all_txt = ''
        self.fields = ["title","areaName","appNoticeTypeName","noticeTypeName","releaseTime","channelId","contentId","migration","projectCode"]
        self.fields_chn = ["名称","区域","公告类型","公告子类型","发布时间","channelId","contentId","migration","projectCode"]
        self.dir_path = dir_path
        
    def main(self):
        try:
            self._first_request()
            print(f"Total Pages: {self.totalPages}, Total Elements: {self.totalElements}")
            for page in range(1, self.totalPages):
                self.params["page"] = page
                self.get_title()
            print(f"Completed fetching all pages. {len(self.content_list)} items collected.")
            self._save_data()
        except Exception as e:
            print(f"An error occurred: {e}")

    def get_title(self):
        response = requests.post(self.url, json=self.params, headers=self.headers, timeout=10)
        response.raise_for_status()
        self._parse_data(response.json())
        print(f"Fetched page {self.params['page']}")
        return response.json()       
    
    def _first_request(self):
        res = self.get_title()
        self.totalPages = res.get("data", {}).get("totalPages", 0)
        self.totalElements = res.get("data", {}).get("totalElements", 0)

    def _parse_data(self, data):
        content = data.get("data", {}).get("content", [])
        content_length = len(content)   
        if (self.params["page"] < self.totalPages -1 and content_length < self.params["size"]) or \
           (self.params["page"] == self.totalPages - 1 and content_length < (self.totalElements % self.params["size"])):
            print(f"Warning: Fewer items {content_length} than expected on page {self.params['page']}")

        jyfwxx_detail = JyfwxxDetail()
        for idx, item in enumerate(content):
            content_item = [item.get(field).replace('\n', '') if field=='title' else item.get(field) for field in self.fields]
            try:
                detail_content_item, detail_txt  = jyfwxx_detail.main(title=item.get("title").replace('\n', ''), contentId=item.get("contentId"), channelId=item.get("channelId"), migration=item.get("migration"), projectCode=item.get("projectCode"), crumb="jtqy", dir_path=self.dir_path)
            except Exception as e:
                print(f"Error fetching detail for  {e}: {content_item[0]} (Content ID: {content_item[6]} Migration: {content_item[7]} Project Code: {content_item[8]})")
                detail_content_item, detail_txt = None, ''
            content_item.append(detail_content_item)
            self.content_list.append(content_item)
            self.all_txt += detail_txt + '<br><hr><br>\n'
            # print(f"Page:{self.params['page']} Index:{idx}: {content_item[0]} (Content ID: {content_item[6]} Migration: {content_item[7]} Project Code: {content_item[8]})")

    def _save_data(self):
        file_path = os.path.join(self.dir_path, f"{self.params['title']}.html")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(self.all_txt)
        print(f"All detail contents saved to {file_path}")

        wb = Workbook()
        ws = wb.active
        ws.append(self.fields_chn + ["详情链接"])
        for item in self.content_list:
            file_name = os.path.basename(item[-1]) if item[-1] else ""
            ws.append(item[:-1] + [file_name])

            link_cell = ws.cell(row=ws.max_row, column=len(self.fields_chn)+1)
            link_cell.hyperlink = item[-1]
            link_cell.style = "Hyperlink"

        ws['N1'] = f"汇总-{self.params['title']}.html"
        ws['N1'].hyperlink = file_path
        ws['N1'].font = Font(color="0000FF", underline="single")  # 蓝色下划线

        excel_path = os.path.join(self.dir_path, f"{self.params['title']}.xlsx")
        wb.save(excel_path)
        print(f"Data saved to Excel file at {excel_path}")

    def save_to_json(self, filename="collected_data.json"):
        """Save collected data to a JSON file"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.content_list, f, ensure_ascii=False, indent=2)
        print(f"Data saved to {filename}")

class JyfwxxDetail():
    def __init__(self):
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36 Edg/143.0.0.0",
        }

    @retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, max=60), before_sleep=lambda retry_state: print(f"{retry_state.outcome.exception()}: ({retry_state.kwargs.get('title')} {retry_state.kwargs.get('contentId')} {retry_state.attempt_number})"))
    def main(self, title, contentId, channelId, migration=None, projectCode=None, crumb="jtqy", dir_path=r"D:\work\交易信息\附件"):
        detail2_url = f"https://rent.szexgrp.com/cms/api/v1/lease/project/detail2?projectId={projectCode}"
        
        res_title = self._request(detail2_url)
        res_title_data = res_title.get("data")
        if res_title.get("code", -1) != 200:
            raise Exception(f"Failed to fetch title type for {title} contentId {contentId}, code: {res_title.get('code', -1)} message: {res_title.get('message', '')}")
        elif not res_title_data:
            detail_url = f"https://www.szexgrp.com/cms/api/v1/trade/content/detail?contentId={contentId}"
            res = self._request(detail_url)            
            file_path, txt = self._parse(res, contentId, dir_path, ori_tile=title, parse_type=1)
        elif res_title_data:
            contentId = res_title_data.get("contentId", contentId)
            detail_project_url = f"https://rent.szexgrp.com/cms/api/v1/lease/project/detailNew?contentId={contentId}"
            res = self._request(detail_project_url)
            file_path, txt = self._parse(res, contentId, dir_path, ori_tile=title, parse_type=2)
        return file_path, txt

    def _request(self, detail_url):
        response = requests.get(detail_url, headers=self.headers, timeout=10)
        response.raise_for_status()
        return response.json()

    def _parse(self, res, contentId, dir_path, ori_tile=None, parse_type=1):
        data = res.get("data", {})
        code = res.get("code", -1)
        message = res.get("message", "")
        if code == 200 and not data:
            print(f"{contentId} No data returned from API.")
            return None, ''
        elif code != 200:
            raise Exception(f"API returned error code: {code} Message: {message}")
        else:
            if parse_type == 1:
                title = data.get("title", '').replace('\n', '')
                releaseTime = data.get("releaseTime", '')
                source = data.get("source", '')
                txt = f'<h1 style="text-align: center; font-family: 楷体, sans-serif; font-size: 36px; color: #e74c3c;">{title}</h1>\n'
                txt += f'<div style="text-align: center;"><span>发布时间：{releaseTime}</span><span>信息来源：{source}</span></div>\n'
                txt += data.get("txt", '')
            elif parse_type == 2:
                title = data.get("projectName", '').replace('\n', '')
                txt = f'<h1 style="text-align: center; font-family: 楷体, sans-serif; font-size: 36px; color: #e74c3c;">{title}</h1>\n'
                for i in data.get('leaseNotices', []):
                    txt += f"<h2 style=\"text-align: center;\">{i.get('noticeTypeName','')}</h2>\n"
                    txt += i.get('noticeContent','') + "\n"

            if ori_tile != title:
                print(f'title different: {contentId} {ori_tile} "!=======================================================================" {title}')

            file_path = os.path.join(dir_path, f"{title}_{contentId}.html")
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(txt)
                # print(f"Saved detail content to {title}_{contentId}.html")
            return file_path, txt

def get_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument('-o', '--out_path', type=str, required=True, help='保存路径')
    parser.add_argument('-s', '--size', type=int, default=50, help='每页数量')
    return parser
        
if __name__ == "__main__":
    parser = get_parser()
    args = parser.parse_args()

    start_time = datetime.now()
    # jyfwxx = Jyfwxx(title="城市更新", size=50, dir_path=r"D:\work\交易信息\附件")
    jyfwxx = Jyfwxx(title="城市更新", size=args.size, dir_path=args.out_path)
    jyfwxx.main()
    end_time = datetime.now()
    print(f"Total time taken: {end_time - start_time} seconds")

    # JyfwxxDetail().main(title="深圳市罗湖区东湖街道布心村水围村城市更新项目一期01-02地块天珺府公寓租赁招租", contentId=2406695, channelId=2855, migration=8, projectCode="93cc5ce3a66b4b04bef6401792553783", dir_path=jyfwxx.dir_path)

    # Save data to JSON file
    # jyfwxx.save_to_json("collected_data.json")
    
    # Optionally print first few items
    # if jyfwxx.content_list:
    #     print("\nFirst item sample:")
    #     print(json.dumps(jyfwxx.content_list[0], ensure_ascii=False, indent=2))